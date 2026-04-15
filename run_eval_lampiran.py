"""Run all 10 evaluation questions and produce an Excel report."""
import os, sys, time
os.environ["GRAPHRAG_STANDALONE"] = "1"

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from utils.langgraph_agent import create_agent

QUESTIONS = [
    {
        "no": 1,
        "soal": "Berapa kebutuhan minimal luas lantai bangunan rumah sakit per tempat tidur menurut peraturan?",
        "jawaban_benar": "Minimal 80 m² per tempat tidur yang dimiliki oleh rumah sakit. Luasan dapat bertambah disesuaikan kapasitas dan kebutuhan pelayanan rumah sakit serta pengembangan ruang-ruang penunjang pelayanan.",
        "sumber": "Bagian C (Kebutuhan Total Luas Lantai Bangunan), Lampiran PMK No. 40 Tahun 2022",
    },
    {
        "no": 2,
        "soal": "Sebutkan 3 zonasi yang digunakan dalam pengelompokan blok bangunan atau area di rumah sakit beserta penjelasannya!",
        "jawaban_benar": "1) Zona merah: area pelayanan pasien penyakit infeksi emerging.\n2) Zona kuning: area pelayanan pasien umum.\n3) Zona hijau: area penunjang dan manajemen.",
        "sumber": "Bagian F (Pola Hubungan Antar Ruang-Ruang) halaman 9, Lampiran PMK No. 40 Tahun 2022",
    },
    {
        "no": 3,
        "soal": "Berapa tinggi minimal langit-langit di ruangan operasi dan di selasar (koridor) rumah sakit?",
        "jawaban_benar": "Tinggi langit-langit di ruangan operasi minimal 3,00 m. Tinggi langit-langit di ruangan minimal 2,80 m dan tinggi di selasar (koridor) minimal 2,40 m.",
        "sumber": "Bagian G poin 3b (Langit-Langit), Lampiran PMK No. 40 Tahun 2022",
    },
    {
        "no": 4,
        "soal": "Berapa kapasitas minimal lahan parkir yang harus disediakan rumah sakit dan apa ketentuannya?",
        "jawaban_benar": "Kapasitas minimal 20% dari luas total bangunan (sudah termasuk jalur sirkulasi kendaraan). Penyediaan lahan parkir tidak boleh mengurangi daerah penghijauan yang telah ditetapkan.",
        "sumber": "Bagian A (Lahan dan Akses Bangunan) poin 4, Lampiran PMK No. 40 Tahun 2022",
    },
    {
        "no": 5,
        "soal": "Apa saja yang termasuk dalam deskripsi umum persyaratan jenjang subkualifikasi tenaga kerja konstruksi?",
        "jawaban_benar": "a. Bertaqwa kepada Tuhan Yang Maha Esa.\nb. Memiliki moral, etika dan kepribadian yang baik di dalam menyelesaikan tugasnya.\nc. Berperan sebagai warga Negara yang bangga dan cinta tanah air serta mendukung perdamaian dunia.\nd. Mampu bekerja sama dan memiliki kepekaan sosial dan kepedulian yang tinggi terhadap masyarakat dan lingkungannya.\ne. Menghargai keanekaragaman budaya, pandangan, kepercayaan, dan agama serta pendapat/temuan original orang lain.\nf. Menjunjung tinggi penegakan hukum serta memiliki semangat untuk mendahulukan kepentingan bangsa serta masyarakat luas.",
        "sumber": "Bagian I (Persyaratan Umum Jenjang Subkualifikasi), Tabel halaman 13, Lampiran SE Dirjen Bina Konstruksi No. 214/SE/Dk/2022",
    },
    {
        "no": 6,
        "soal": "Jelaskan kemampuan, pengetahuan, dan tanggung jawab yang harus dimiliki oleh tenaga kerja Terampil Kelas 3!",
        "jawaban_benar": "Kemampuan: Mampu melaksanakan tugas sederhana, terbatas, bersifat rutin, dengan menggunakan alat, aturan, dan proses yang telah ditetapkan, serta di bawah bimbingan, pengawasan, dan tanggung jawab atasannya.\nPengetahuan: Memiliki pengetahuan faktual.\nTanggung jawab: Bertanggung jawab atas pekerjaan sendiri dan tidak bertanggung jawab atas pekerjaan orang lain.",
        "sumber": "Bagian I (Persyaratan Umum Jenjang Subkualifikasi), Tabel halaman 13, Lampiran SE Dirjen Bina Konstruksi No. 214/SE/Dk/2022",
    },
    {
        "no": 7,
        "soal": "Apa persyaratan pendidikan dan pengalaman minimal untuk menjadi Ahli Muda di klasifikasi Arsitektural (Arsitek)?",
        "jawaban_benar": "Ahli Muda:\n• Pendidikan Profesi minimal 0 tahun pengalaman, atau\n• S1/S1 Terapan/D4 Terapan minimal 2 tahun pengalaman.\nProgram studi: Arsitektur/Teknik Arsitektur + PPAR atau Arsitektur/Teknik Arsitektur + STRA.\nAcuan: SKKNI 196-2021.",
        "sumber": "Bagian II (Klasifikasi dan Subklasifikasi Tenaga Kerja Ahli Konstruksi), Lampiran SE Dirjen Bina Konstruksi No. 214/SE/Dk/2022",
    },
    {
        "no": 8,
        "soal": "Apa kode jabatan kerja baru untuk Asisten Arsitek pada klasifikasi Arsitektur dan apa acuan SKKNI-nya?",
        "jawaban_benar": "Kode: ARS.01.001.7, Jenjang 7.\nAcuan: SKKNI 196-2021.\nPendidikan: Arsitektur/Teknik Arsitektur + PPAR.\nKetentuan Persyaratan SKK Asesor: Seluruh SKK Konstruksi pada subklasifikasi Arsitektural (Jenjang 8, 9).",
        "sumber": "Tabel halaman 5, Lampiran Keputusan Dirjen Bina Konstruksi No. 12.1/KPTS/Dk/2022",
    },
    {
        "no": 9,
        "soal": "Apa persyaratan pendidikan/program studi untuk jabatan Ahli Penilai Kelaikan Bangunan Gedung (Aspek Arsitektur dan Tata Ruang Luar)?",
        "jawaban_benar": "Pendidikan/Program Studi: Teknik Sipil; Arsitektur/Teknik Arsitektur.\nJenjang: 9, Kode: SIP.01.010.9.\nAcuan: SKKNI 113-2015.\nKetentuan SKK Asesor: Seluruh SKK Konstruksi pada subklasifikasi Gedung atau Arsitektural (Jenjang 9).",
        "sumber": "Tabel halaman 6, Lampiran Keputusan Dirjen Bina Konstruksi No. 12.1/KPTS/Dk/2022",
    },
    {
        "no": 10,
        "soal": "Apa acuan SKKNI dan persyaratan pendidikan untuk jabatan Ahli Madya Perencanaan Jembatan Rangka Baja?",
        "jawaban_benar": "Acuan: SKKNI 130-2015.\nPendidikan: Teknik Sipil.\nJenjang: 8, Kode: SIP.04.005.8.\nKetentuan SKK Asesor: Seluruh SKK Konstruksi pada subklasifikasi Jembatan (Jenjang 9).",
        "sumber": "Tabel halaman 7, Lampiran Keputusan Dirjen Bina Konstruksi No. 12.1/KPTS/Dk/2022",
    },
]

def run_question(agent, q):
    """Run a single question and return the answer + metadata."""
    result = agent.invoke({
        "query": q["soal"],
        "route": "",
        "primary_doc_ids": [],
        "context_docs": {},
        "relationship_context": "",
        "answer": "",
        "logs": [],
        "narratives": [],
        "chat_history": [],
        "summary": "",
        "user_context": "",
    })
    lamp_chunks = sum(
        1 for info in result.get("context_docs", {}).values()
        for ch in info.get("chunks", [])
        if ch.get("scope") == "neo4j-lampiran"
    )
    total_chunks = sum(
        len(info.get("chunks", []))
        for info in result.get("context_docs", {}).values()
    )
    return {
        "answer": result.get("answer", ""),
        "route": result.get("route", ""),
        "docs": result.get("primary_doc_ids", []),
        "total_chunks": total_chunks,
        "lamp_chunks": lamp_chunks,
    }


def main():
    agent = create_agent()
    results = []

    for q in QUESTIONS:
        print(f"\n{'='*60}")
        print(f"Soal {q['no']}: {q['soal'][:80]}...")
        t0 = time.time()
        res = run_question(agent, q)
        elapsed = time.time() - t0
        print(f"  Route: {res['route']}, Docs: {res['docs'][:3]}")
        print(f"  Chunks: {res['total_chunks']} total, {res['lamp_chunks']} lampiran")
        print(f"  Time: {elapsed:.1f}s")
        print(f"  Answer preview: {res['answer'][:200]}...")
        results.append({**q, **res})

    # Write Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evaluasi Lampiran"

    # Headers
    headers = ["No", "Soal", "Jawaban Chatbot", "Jawaban Benar", "Sumber", "Route", "Docs", "Lampiran Chunks"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Data rows
    wrap_align = Alignment(vertical="top", wrap_text=True)
    for i, r in enumerate(results, 2):
        row_data = [
            r["no"],
            r["soal"],
            r["answer"],
            r["jawaban_benar"],
            r["sumber"],
            r["route"],
            ", ".join(r["docs"][:5]),
            f"{r['lamp_chunks']}/{r['total_chunks']}",
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = wrap_align
            cell.border = thin_border

    # Column widths
    widths = [5, 40, 60, 50, 35, 10, 40, 15]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    out_path = "output/evaluasi_lampiran_results.xlsx"
    wb.save(out_path)
    print(f"\n{'='*60}")
    print(f"Results saved to {out_path}")


if __name__ == "__main__":
    main()
