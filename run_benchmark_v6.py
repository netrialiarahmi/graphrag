#!/usr/bin/env python3
"""
Benchmark v6 — GraphRAG (VDB 10 + Neo4j 3-child) + GPT-4.1 Re-Rank.

Pipeline per question:
  1. Embed question via HuggingFace
  2. Pinecone semantic_search(top_k=100) → 10 unique docs
  3. Neo4j: for each of the 10 docs, get ≤3 related docs (CITES/HIGHER)
  4. Build content snippets for all candidate docs
  5. GPT-4.1 re-rank: score 0-10 per document → sort descending
  6. Evaluate @5, @10, @20

Output naming convention:
  detail/   GraphRAG+GPT_ReRank-{QA label}.csv        (raw retrieval per Q)
  detail/   GraphRAG+GPT_ReRank-{QA label}-metrics.csv (per-Q metrics)
  detail/   GraphRAG+GPT_ReRank-{QA label}-metrics.xlsx
  recap/    GraphRAG+GPT_ReRank-{QA label}.csv         (aggregate summary)

Usage:
    python run_benchmark_v6.py                         # all .xlsx in benchmark/
    python run_benchmark_v6.py path/to/file.xlsx       # single file
"""

import argparse
import csv
import glob
import json
import math
import os
import sys
import time

os.environ["GRAPHRAG_STANDALONE"] = "1"
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from dotenv import load_dotenv
load_dotenv()

from utils import llm_stance, neo4j_client, pinecone_client
from utils.benchmark_helpers import (
    extract_documents,
    get_correct_doc_id,
    get_unique_doc_ids,
    extract_doc_ids_from_question,
    build_doc_id_aliases,
    match_with_aliases,
)

# ── Tuning constants ──────────────────────────────────────────────────────────
VDB_TOP_K        = 100   # Raw Pinecone results
VDB_MAX_DOCS     = 10    # Unique VDB doc cap
NEO4J_CHILDREN   = 3     # Max related docs per VDB doc
GPT_MAX_TOKENS   = 1500  # Max tokens for GPT re-rank response
GPT_TEMPERATURE  = 0.1

# ── Paths ─────────────────────────────────────────────────────────────────────
ROOT_DIR      = os.path.dirname(os.path.abspath(__file__))
BENCHMARK_DIR = os.path.join(ROOT_DIR, "benchmark")
DETAIL_CSV_DIR  = os.path.join(ROOT_DIR, "output", "retrieval", "detail", "csv")
DETAIL_XLSX_DIR = os.path.join(ROOT_DIR, "output", "retrieval", "detail", "xlsx")
RECAP_DIR       = os.path.join(ROOT_DIR, "output", "retrieval", "recap")

METHOD_NAME = "GraphRAG+GPT_ReRank"

QA_LABELS = {
    "QA 100 (test-all-sector)": "QA 100",
    "govnetic_qa_complete_50 (business)": "QA Business",
}

CUTOFFS = [5, 10, 20]


# ── Question parser ──────────────────────────────────────────────────────────

def _parse_questions(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    questions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        q_no = str(row[0])
        q_text = str(row[1]) if row[1] else ""
        evidence = str(row[3]) if len(row) > 3 and row[3] else ""
        parsed_docs = extract_documents(evidence)
        gt_doc_ids, all_candidates = set(), set()
        for d in parsed_docs:
            did = get_correct_doc_id(d)
            if did:
                gt_doc_ids.add(did)
                all_candidates.add(did)
            for c in d.get("candidates", []):
                all_candidates.add(c)
        q_doc_ids = extract_doc_ids_from_question(q_text)
        gt_doc_ids.update(q_doc_ids)
        all_candidates.update(q_doc_ids)
        if q_text:
            questions.append({
                "no": q_no, "question": q_text,
                "gt_doc_ids": gt_doc_ids,
                "all_candidates": all_candidates,
                "q_doc_ids": q_doc_ids,
            })
    wb.close()
    return questions


# ── Build doc summaries for GPT re-ranking ────────────────────────────────────

def _build_doc_summary(doc_id: str, semantic_chunks: dict[str, list[dict]],
                       neo4j_ok: bool) -> str:
    """Get a content snippet for a doc_id (for GPT prompt)."""
    # Try VDB semantic chunks first
    if doc_id in semantic_chunks:
        return semantic_chunks[doc_id][0].get("content", "")[:500]

    # Try fetching from Pinecone by doc_id
    try:
        chunks = pinecone_client.fetch_by_doc_id(doc_id, top_k=3)
        if chunks:
            return chunks[0].get("content", "")[:500]
    except Exception:
        pass

    # Fallback: Neo4j pasal content
    if neo4j_ok:
        try:
            detail = neo4j_client.get_document_detail(doc_id)
            pasals = detail.get("pasals", [])
            if pasals:
                texts = [p.get("content", "") or p.get("name", "") for p in pasals[:3]]
                return " ".join(t for t in texts if t)[:500]
        except Exception:
            pass

    return ""


# ── GPT-4.1 re-rank (reuse existing llm_stance pattern) ──────────────────────

def gpt_rerank(query: str, doc_summaries: dict[str, str]) -> list[str]:
    """Send all candidate docs to GPT-4.1 for relevance scoring.

    Returns list of doc_ids sorted by GPT score descending.
    """
    if not doc_summaries:
        return []

    client = llm_stance.get_llm_client()

    summary_parts = []
    for did, text in doc_summaries.items():
        snippet = text[:400] if text else "(kosong)"
        summary_parts.append(f"DOC_ID: {did}\nKonten: {snippet}")
    summaries_str = "\n\n".join(summary_parts)

    system_prompt = (
        "Kamu adalah pakar hukum Indonesia. Berikan skor relevansi 0-10 "
        "untuk setiap dokumen terhadap pertanyaan yang diberikan.\n\n"
        "Format output HARUS berupa JSON array, contoh:\n"
        '[{"doc_id": "UU-NASIONAL-40-2007", "score": 9}, '
        '{"doc_id": "PP-NASIONAL-16-2021", "score": 1}]\n\n'
        "Skor:\n"
        "- 8-10: Sangat relevan, kemungkinan besar memuat jawaban\n"
        "- 5-7: Cukup relevan, mungkin memuat konteks pendukung\n"
        "- 2-4: Sedikit relevan, hubungan tidak langsung\n"
        "- 0-1: Tidak relevan sama sekali\n\n"
        "Output HANYA JSON array, tanpa teks lain."
    )

    try:
        response = client.chat.completions.create(
            model=llm_stance.LLM_MODEL,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": f"Pertanyaan: {query}\n\nDokumen:\n{summaries_str}"},
            ],
            max_tokens=GPT_MAX_TOKENS,
            temperature=GPT_TEMPERATURE,
        )
        raw = response.choices[0].message.content or "[]"
        raw = raw.strip()
        if raw.startswith("```"):
            raw = raw.split("\n", 1)[-1].rsplit("```", 1)[0].strip()

        scored = json.loads(raw)
        results = []
        seen = set()
        for item in scored:
            did = item.get("doc_id", "")
            score = float(item.get("score", 0))
            if did and did not in seen:
                results.append((did, score))
                seen.add(did)
        results.sort(key=lambda x: x[1], reverse=True)

        ranked_ids = [did for did, _ in results]

        # Ensure any docs GPT missed are appended at the end
        for did in doc_summaries:
            if did not in seen:
                ranked_ids.append(did)

        return ranked_ids

    except Exception as e:
        print(f" [WARN] GPT rerank failed: {e}")
        return list(doc_summaries.keys())


# ── Metric functions ──────────────────────────────────────────────────────────

def _is_relevant(doc_id, gt, aliases):
    if doc_id in gt:
        return True
    return bool(aliases.get(doc_id, {doc_id}) & gt)

def _matched_gt_at_k(ranked, gt, aliases, k):
    matched = set()
    top_k = set(ranked[:k])
    for gt_id in gt:
        if gt_id in top_k:
            matched.add(gt_id)
            continue
        if top_k & aliases.get(gt_id, {gt_id}):
            matched.add(gt_id)
    return matched

def recall_at_k(ranked, gt, aliases, k):
    if not gt: return 0.0
    return len(_matched_gt_at_k(ranked, gt, aliases, k)) / len(gt)

def precision_at_k(ranked, gt, aliases, k):
    if k == 0: return 0.0
    n_rel = sum(1 for d in ranked[:k] if _is_relevant(d, gt, aliases))
    return n_rel / k

def f1_at_k(p, r):
    return 2 * p * r / (p + r) if (p + r) > 0 else 0.0

def hit_rate_at_k(ranked, gt, aliases, k):
    return 1.0 if _matched_gt_at_k(ranked, gt, aliases, k) else 0.0

def reciprocal_rank(ranked, gt, aliases):
    for i, doc in enumerate(ranked):
        if _is_relevant(doc, gt, aliases):
            return 1.0 / (i + 1)
    return 0.0

def average_precision_at_k(ranked, gt, aliases, k):
    if not gt: return 0.0
    top_k = ranked[:k]
    n_rel, sum_p = 0, 0.0
    for i, doc in enumerate(top_k):
        if _is_relevant(doc, gt, aliases):
            n_rel += 1
            sum_p += n_rel / (i + 1)
    normaliser = min(k, len(gt))
    return sum_p / normaliser if normaliser > 0 else 0.0

def ndcg_at_k(ranked, gt, aliases, k):
    if not gt: return 0.0
    top_k = ranked[:k]
    dcg = sum(1.0 / math.log2(i + 2) for i, d in enumerate(top_k) if _is_relevant(d, gt, aliases))
    ideal = min(k, len(gt))
    idcg = sum(1.0 / math.log2(i + 2) for i in range(ideal))
    return dcg / idcg if idcg > 0 else 0.0


# ── Compute metrics for the re-ranked list ────────────────────────────────────

def compute_metrics(results: list[dict]) -> tuple[list[dict], dict]:
    detail_rows = []
    accum = {}

    for row in results:
        gt_str = row.get("GT_Doc_IDs", "")
        gt = set(d.strip() for d in gt_str.split(",") if d.strip()) if gt_str else set()
        ranked_str = row.get("Dok_Ranked", "")
        if not ranked_str or ranked_str.startswith("Error"):
            continue
        ranked = [d.strip() for d in ranked_str.split(",") if d.strip()]
        if not gt:
            continue

        q_ids_str = row.get("GT_From_Question", "")
        q_ids = set(d.strip() for d in q_ids_str.split(",") if d.strip()) if q_ids_str else set()
        all_ids = gt | set(ranked) | q_ids
        aliases = build_doc_id_aliases(all_ids)

        detail = {
            "No": row.get("No", ""),
            "Pertanyaan": row.get("Pertanyaan", "")[:150],
            "GT_Total": len(gt),
        }

        mrr = reciprocal_rank(ranked, gt, aliases)
        detail["MRR"] = f"{mrr:.4f}"
        accum.setdefault("MRR", []).append(mrr)

        for k in CUTOFFS:
            detail[f"Dok@{k}"] = ", ".join(ranked[:k])

        for k in CUTOFFS:
            r = recall_at_k(ranked, gt, aliases, k)
            p = precision_at_k(ranked, gt, aliases, k)
            f = f1_at_k(p, r)
            h = hit_rate_at_k(ranked, gt, aliases, k)
            ap = average_precision_at_k(ranked, gt, aliases, k)
            nd = ndcg_at_k(ranked, gt, aliases, k)

            for name, val in [("Recall", r), ("Precision", p), ("F1", f),
                              ("HitRate", h), ("AP", ap), ("NDCG", nd)]:
                detail[f"{name}@{k}"] = f"{val:.4f}"
                accum.setdefault(f"{name}@{k}", []).append(val)

        detail_rows.append(detail)

    n = len(detail_rows)
    summary = {"Scored_Questions": n}
    for key, vals in accum.items():
        summary[f"Avg_{key}"] = sum(vals) / len(vals) if vals else 0.0

    return detail_rows, summary


# ── Writers ───────────────────────────────────────────────────────────────────

def _detail_columns():
    cols = ["No", "Pertanyaan", "GT_Total", "MRR"]
    for k in CUTOFFS:
        cols.append(f"Dok@{k}")
    for k in CUTOFFS:
        for m in ["Recall", "Precision", "F1", "HitRate", "AP", "NDCG"]:
            cols.append(f"{m}@{k}")
    return cols


def write_detail_csv(path, rows):
    cols = _detail_columns()
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def write_recap_csv(path, summary):
    rows_out = [{"Metric": "Scored_Questions", "Value": summary["Scored_Questions"]}]
    rows_out.append({"Metric": "MRR", "Value": f"{summary.get('Avg_MRR', 0):.4f}"})
    for k in CUTOFFS:
        for m in ["Recall", "Precision", "F1", "HitRate", "AP", "NDCG"]:
            v = summary.get(f"Avg_{m}@{k}", 0)
            rows_out.append({"Metric": f"{m}@{k}", "Value": f"{v:.4f}"})
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["Metric", "Value"])
        w.writeheader()
        w.writerows(rows_out)


def write_xlsx(csv_path, xlsx_path):
    def fmt(val):
        try:
            fv = float(val)
            s = f"{fv:.4f}".rstrip('0')
            if s.endswith('.'):
                s += '0'
            return s.replace('.', ',')
        except ValueError:
            return val
    wb = openpyxl.Workbook()
    ws = wb.active
    with open(csv_path, newline='', encoding='utf-8') as f:
        reader = csv.reader(f)
        header = next(reader)
        ws.append(header)
        for row in reader:
            ws.append([fmt(c) for c in row])
    wb.save(xlsx_path)


def print_summary(label, summary):
    n = summary["Scored_Questions"]
    print(f"\n  {'='*60}")
    print(f"  {label}  ({n} scored)")
    print(f"  {'='*60}")
    print(f"  {'Metric':<16} {'Value':>10}")
    print(f"  {'─'*16} {'─'*10}")
    print(f"  {'MRR':<16} {summary.get('Avg_MRR',0):>10.4f}")
    for k in CUTOFFS:
        print(f"  {'─'*16} {'─'*10}")
        for m in ["Recall", "Precision", "F1", "HitRate", "AP", "NDCG"]:
            v = summary.get(f"Avg_{m}@{k}", 0)
            print(f"  {m+'@'+str(k):<16} {v:>10.4f}")
    print(f"  {'='*60}")


# ── Main pipeline ────────────────────────────────────────────────────────────

def process_xlsx(xlsx_path: str):
    base = os.path.splitext(os.path.basename(xlsx_path))[0]
    qa_label = QA_LABELS.get(base, base)

    print(f"\n{'='*70}")
    print(f"  Benchmark v6: {METHOD_NAME} — {qa_label}")
    print(f"  VDB={VDB_MAX_DOCS} docs, Neo4j={NEO4J_CHILDREN} children/doc, GPT rerank")
    print(f"{'='*70}")

    # Connectivity
    if not pinecone_client.test_connection():
        print("  ERROR: Pinecone not connected!"); return
    print("  Pinecone ✓")

    print("  Testing HF embedding...")
    hf_ok = False
    for attempt in range(5):
        if llm_stance.test_hf_connection():
            hf_ok = True; break
        print(f"    Attempt {attempt+1}/5 failed, retrying in 15s...")
        time.sleep(15)
    if not hf_ok:
        print("  ERROR: HF embedding not available!"); return
    print("  HuggingFace ✓")

    try:
        neo4j_ok = neo4j_client.test_connection()
    except Exception:
        neo4j_ok = False
    print(f"  Neo4j {'✓' if neo4j_ok else '✗ (no graph expansion)'}")

    # Test GPT
    try:
        gpt_ok = llm_stance.test_connection()
    except Exception:
        gpt_ok = False
    if not gpt_ok:
        print("  ERROR: GPT (OpenRouter) not connected!"); return
    print(f"  GPT ({llm_stance.LLM_MODEL}) ✓")

    questions = _parse_questions(xlsx_path)
    print(f"  Questions: {len(questions)}\n")

    # ── Process each question ─────────────────────────────────────────────
    raw_results = []
    t0 = time.time()

    for idx, q in enumerate(questions):
        pct = (idx + 1) / len(questions) * 100
        short_q = q["question"][:60]
        print(f"  [{idx+1}/{len(questions)}] ({pct:.0f}%) {short_q}…", end="", flush=True)

        gt = q["gt_doc_ids"]
        q_injected = q["q_doc_ids"]

        if not gt:
            raw_results.append({
                "No": q["no"], "Pertanyaan": q["question"][:200],
                "GT_Total": 0, "GT_Doc_IDs": "", "GT_From_Question": "",
                "Dok_VDB": "", "Dok_GraphRAG": "", "Dok_Ranked": "",
            })
            print(" SKIP (no GT)")
            continue

        try:
            # Step 1: VDB retrieval — 10 unique docs
            q_embedding = llm_stance.get_embedding(q["question"])
            vdb_raw = pinecone_client.semantic_search(
                query_embedding=q_embedding, top_k=VDB_TOP_K
            )
            vdb_doc_ids = get_unique_doc_ids(vdb_raw, VDB_MAX_DOCS)

            # Build semantic chunks map for summary building
            semantic_chunks: dict[str, list[dict]] = {}
            for hit in vdb_raw:
                did = hit.get("doc_id", "")
                if did:
                    semantic_chunks.setdefault(did, []).append(hit)

            # Step 2: Neo4j expansion — ≤3 children per VDB doc
            graphrag_set: dict[str, None] = {}
            for did in vdb_doc_ids:
                graphrag_set[did] = None

            if neo4j_ok:
                for did in vdb_doc_ids:
                    try:
                        related = neo4j_client.get_related_documents(
                            did, limit=NEO4J_CHILDREN
                        )
                        for rdoc in related:
                            rdid = rdoc.get("doc_id", "")
                            if rdid and rdid not in graphrag_set:
                                graphrag_set[rdid] = None
                    except Exception:
                        pass

            all_candidate_ids = list(graphrag_set.keys())

            # Step 3: Build summaries for GPT
            doc_summaries: dict[str, str] = {}
            for did in all_candidate_ids:
                doc_summaries[did] = _build_doc_summary(did, semantic_chunks, neo4j_ok)

            # Step 4: GPT-4.1 re-rank
            ranked_ids = gpt_rerank(q["question"], doc_summaries)

            raw_results.append({
                "No": q["no"],
                "Pertanyaan": q["question"][:200],
                "GT_Total": len(gt),
                "GT_Doc_IDs": ", ".join(sorted(gt)),
                "GT_From_Question": ", ".join(sorted(q_injected)) if q_injected else "",
                "Dok_VDB": ", ".join(vdb_doc_ids),
                "Dok_GraphRAG": ", ".join(all_candidate_ids),
                "Dok_Ranked": ", ".join(ranked_ids),
            })
            n_total = len(all_candidate_ids)
            n_neo4j = n_total - len(vdb_doc_ids)
            print(f" ✓ VDB={len(vdb_doc_ids)} +Neo4j={n_neo4j} → {n_total} candidates → GPT ranked")

        except Exception as e:
            raw_results.append({
                "No": q["no"], "Pertanyaan": q["question"][:200],
                "GT_Total": len(gt),
                "GT_Doc_IDs": ", ".join(sorted(gt)),
                "GT_From_Question": ", ".join(sorted(q_injected)) if q_injected else "",
                "Dok_VDB": f"Error: {e}", "Dok_GraphRAG": "", "Dok_Ranked": "",
            })
            print(f" ✗ {e}")

    elapsed = time.time() - t0
    print(f"\n  Pipeline done in {elapsed:.1f}s")

    # ── Write outputs ─────────────────────────────────────────────────────
    os.makedirs(DETAIL_DIR, exist_ok=True)
    os.makedirs(RECAP_DIR, exist_ok=True)

    label = f"{METHOD_NAME}-{qa_label}"

    # Raw retrieval detail
    raw_cols = ["No", "Pertanyaan", "GT_Total", "GT_Doc_IDs", "GT_From_Question",
                "Dok_VDB", "Dok_GraphRAG", "Dok_Ranked"]
    raw_path = os.path.join(DETAIL_DIR, f"{label}.csv")
    with open(raw_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=raw_cols, extrasaction="ignore")
        w.writeheader()
        w.writerows(raw_results)
    print(f"  → {raw_path}")

    # Metrics
    detail_rows, summary = compute_metrics(raw_results)

    metrics_csv = os.path.join(DETAIL_CSV_DIR, f"{label}-metrics.csv")
    write_detail_csv(metrics_csv, detail_rows)
    print(f"  → {metrics_csv}")

    metrics_xlsx = os.path.join(DETAIL_XLSX_DIR, f"{label}-metrics.xlsx")
    write_xlsx(metrics_csv, metrics_xlsx)
    print(f"  → {metrics_xlsx}")

    recap_csv = os.path.join(RECAP_DIR, f"{label}.csv")
    write_recap_csv(recap_csv, summary)
    print(f"  → {recap_csv}")

    print_summary(label, summary)


# ── Entry point ──────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Benchmark v6 — GraphRAG + GPT ReRank")
    parser.add_argument("xlsx", nargs="?", default=None)
    args = parser.parse_args()

    if args.xlsx:
        files = [args.xlsx]
    else:
        files = sorted(glob.glob(os.path.join(BENCHMARK_DIR, "*.xlsx")))
        if not files:
            print("No .xlsx files in benchmark/"); sys.exit(1)

    print(f"Processing {len(files)} file(s)")
    for f in files:
        process_xlsx(f)
    print("\n✅ All done!")


if __name__ == "__main__":
    main()
