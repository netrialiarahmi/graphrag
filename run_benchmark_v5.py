#!/usr/bin/env python3
"""
Run IR Benchmark v5 — VDB retrieval (20 docs) + Neo4j graph re-ranking.

Pipeline per question:
  1. Embed question via HuggingFace
  2. Retrieve 20 unique docs from Pinecone VDB (top_k=100 → deduplicate)
  3. Re-rank the 20 docs using Neo4j graph signal:
       - For each doc, count CITES/HIGHER edges to other docs in the set
       - combined_score = vdb_score + α × graph_degree
       - Sort descending by combined_score
  4. Output ranked lists for both VDB (original order) and GraphRAG (re-ranked)
  5. Compute metrics @5, @10, @20

Output naming: Method-QA Name-(detail/recap)
  output/retrieval/detailed retrieval/VDB-{qa_label}-detail.csv
  output/retrieval/detailed retrieval/GraphRAG-{qa_label}-detail.csv
  output/retrieval/metrics/VDB-{qa_label}-metrics_detail.csv
  output/retrieval/metrics/VDB-{qa_label}-metrics_recap.csv
  output/retrieval/metrics/GraphRAG-{qa_label}-metrics_detail.csv
  output/retrieval/metrics/GraphRAG-{qa_label}-metrics_recap.csv

Usage:
    python run_benchmark_v5.py                              # all .xlsx
    python run_benchmark_v5.py path/to/file.xlsx            # single file
"""

import argparse
import csv
import glob
import math
import os
import sys
import time

os.environ["GRAPHRAG_STANDALONE"] = "1"
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from dotenv import load_dotenv
load_dotenv()

from utils import llm_stance, neo4j_client, pinecone_client
from utils.benchmark_helpers import (
    extract_documents,
    get_correct_doc_id,
    get_unique_doc_ids,
    extract_doc_ids_from_question,
    build_doc_id_aliases,
    match_with_aliases,
)

# ── Tuning constants ──────────────────────────────────────────────────────────
VDB_TOP_K = 100           # Raw results from Pinecone
VDB_MAX_DOCS = 20         # Unique doc cap for VDB
RERANK_ALPHA = 0.1        # Weight for graph signal in combined score

# ── Paths ─────────────────────────────────────────────────────────────────────
ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
BENCHMARK_DIR = os.path.join(ROOT_DIR, "benchmark")
DETAIL_CSV_DIR  = os.path.join(ROOT_DIR, "output", "retrieval", "detail", "csv")
DETAIL_XLSX_DIR = os.path.join(ROOT_DIR, "output", "retrieval", "detail", "xlsx")
RECAP_DIR       = os.path.join(ROOT_DIR, "output", "retrieval", "recap")

# ── QA label map (xlsx basename → short label) ───────────────────────────────
QA_LABELS = {
    "QA 100 (test-all-sector)": "QA 100",
    "govnetic_qa_complete_50 (business)": "QA Business",
}

CUTOFFS = [5, 10, 20]


# ── Question parser ──────────────────────────────────────────────────────────

def _parse_questions(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    questions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        q_no = str(row[0])
        q_text = str(row[1]) if row[1] else ""
        evidence = str(row[3]) if len(row) > 3 and row[3] else ""
        parsed_docs = extract_documents(evidence)
        gt_doc_ids, all_candidates = set(), set()
        for d in parsed_docs:
            did = get_correct_doc_id(d)
            if did:
                gt_doc_ids.add(did)
                all_candidates.add(did)
            for c in d.get("candidates", []):
                all_candidates.add(c)
        q_doc_ids = extract_doc_ids_from_question(q_text)
        gt_doc_ids.update(q_doc_ids)
        all_candidates.update(q_doc_ids)
        if q_text:
            questions.append({
                "no": q_no, "question": q_text,
                "gt_doc_ids": gt_doc_ids,
                "all_candidates": all_candidates,
                "q_doc_ids": q_doc_ids,
            })
    wb.close()
    return questions


# ── Re-ranking with Neo4j graph signal ────────────────────────────────────────

def rerank_with_graph(vdb_docs_with_scores: list[tuple[str, float]],
                      neo4j_ok: bool) -> list[str]:
    """Re-rank VDB documents using graph connectivity.

    Args:
        vdb_docs_with_scores: [(doc_id, vdb_score), ...] in VDB rank order
        neo4j_ok: whether Neo4j is available

    Returns:
        Re-ranked list of doc_ids (best first).

    Strategy:
        For each doc in the set, count how many CITES/HIGHER edges connect it
        to OTHER docs in the same set.  Normalize graph_degree to [0, 1].
        combined = vdb_score_normalized + α × graph_degree_normalized
        Sort descending.
    """
    if not vdb_docs_with_scores:
        return []

    doc_ids = [d[0] for d in vdb_docs_with_scores]

    # Normalize VDB scores to [0, 1]
    scores = [d[1] for d in vdb_docs_with_scores]
    max_s, min_s = max(scores), min(scores)
    score_range = max_s - min_s if max_s > min_s else 1.0
    vdb_norm = {did: (s - min_s) / score_range for did, s in vdb_docs_with_scores}

    # Graph degree (edges between docs in the set)
    graph_degree = {did: 0 for did in doc_ids}
    if neo4j_ok:
        try:
            edges_data = neo4j_client.get_edges_between(doc_ids)
            for edge in edges_data.get("edges", []):
                src = edge.get("source_id", "")
                tgt = edge.get("target_id", "")
                if src in graph_degree:
                    graph_degree[src] += 1
                if tgt in graph_degree:
                    graph_degree[tgt] += 1
        except Exception as e:
            print(f" [WARN] Neo4j edge query failed: {e}")

    # Normalize graph degree to [0, 1]
    max_deg = max(graph_degree.values()) if graph_degree else 0
    if max_deg > 0:
        graph_norm = {did: deg / max_deg for did, deg in graph_degree.items()}
    else:
        graph_norm = {did: 0.0 for did in doc_ids}

    # Combined score
    combined = []
    for did in doc_ids:
        c = vdb_norm.get(did, 0) + RERANK_ALPHA * graph_norm.get(did, 0)
        combined.append((did, c))

    # Sort descending by combined score
    combined.sort(key=lambda x: x[1], reverse=True)
    return [did for did, _ in combined]


# ── Metric functions (same as run_metrics_at_k.py) ───────────────────────────

def _is_relevant(doc_id, gt, aliases):
    if doc_id in gt:
        return True
    return bool(aliases.get(doc_id, {doc_id}) & gt)

def _matched_gt_at_k(ranked, gt, aliases, k):
    matched = set()
    top_k = set(ranked[:k])
    for gt_id in gt:
        if gt_id in top_k:
            matched.add(gt_id)
            continue
        if top_k & aliases.get(gt_id, {gt_id}):
            matched.add(gt_id)
    return matched

def recall_at_k(ranked, gt, aliases, k):
    if not gt: return 0.0
    return len(_matched_gt_at_k(ranked, gt, aliases, k)) / len(gt)

def precision_at_k(ranked, gt, aliases, k):
    if k == 0: return 0.0
    top_k = ranked[:k]
    n_rel = sum(1 for d in top_k if _is_relevant(d, gt, aliases))
    return n_rel / k

def f1_at_k(p, r):
    return 2 * p * r / (p + r) if (p + r) > 0 else 0.0

def hit_rate_at_k(ranked, gt, aliases, k):
    return 1.0 if _matched_gt_at_k(ranked, gt, aliases, k) else 0.0

def reciprocal_rank(ranked, gt, aliases):
    for i, doc in enumerate(ranked):
        if _is_relevant(doc, gt, aliases):
            return 1.0 / (i + 1)
    return 0.0

def average_precision_at_k(ranked, gt, aliases, k):
    if not gt: return 0.0
    top_k = ranked[:k]
    n_rel, sum_p = 0, 0.0
    for i, doc in enumerate(top_k):
        if _is_relevant(doc, gt, aliases):
            n_rel += 1
            sum_p += n_rel / (i + 1)
    normaliser = min(k, len(gt))
    return sum_p / normaliser if normaliser > 0 else 0.0

def ndcg_at_k(ranked, gt, aliases, k):
    if not gt: return 0.0
    top_k = ranked[:k]
    dcg = sum(1.0 / math.log2(i + 2) for i, d in enumerate(top_k) if _is_relevant(d, gt, aliases))
    ideal_count = min(k, len(gt))
    idcg = sum(1.0 / math.log2(i + 2) for i in range(ideal_count))
    return dcg / idcg if idcg > 0 else 0.0


# ── Compute metrics for a single method ──────────────────────────────────────

def compute_single_method_metrics(results: list[dict], doc_key: str) -> tuple[list[dict], dict]:
    """Compute metrics @K for one method (VDB or GraphRAG).

    Args:
        results: list of per-question result dicts
        doc_key: column name for ranked docs (e.g. "Dok_VDB" or "Dok_GraphRAG")

    Returns:
        (detail_rows, summary_dict)
    """
    detail_rows = []
    accum = {}

    for row in results:
        gt_str = row.get("GT_Doc_IDs", "")
        gt = set(d.strip() for d in gt_str.split(",") if d.strip()) if gt_str else set()
        ranked_str = row.get(doc_key, "")
        if not ranked_str or ranked_str.startswith("Error"):
            continue
        ranked = [d.strip() for d in ranked_str.split(",") if d.strip()]
        if not gt:
            continue

        q_ids_str = row.get("GT_From_Question", "")
        q_ids = set(d.strip() for d in q_ids_str.split(",") if d.strip()) if q_ids_str else set()
        all_ids = gt | set(ranked) | q_ids
        aliases = build_doc_id_aliases(all_ids)

        detail = {
            "No": row.get("No", ""),
            "Pertanyaan": row.get("Pertanyaan", "")[:150],
            "GT_Total": len(gt),
        }

        # MRR
        mrr = reciprocal_rank(ranked, gt, aliases)
        detail["MRR"] = f"{mrr:.4f}"
        accum.setdefault("MRR", []).append(mrr)

        # Docs at each cutoff
        for k in CUTOFFS:
            detail[f"Dok@{k}"] = ", ".join(ranked[:k])

        # Metrics per cutoff
        for k in CUTOFFS:
            r = recall_at_k(ranked, gt, aliases, k)
            p = precision_at_k(ranked, gt, aliases, k)
            f = f1_at_k(p, r)
            h = hit_rate_at_k(ranked, gt, aliases, k)
            ap = average_precision_at_k(ranked, gt, aliases, k)
            nd = ndcg_at_k(ranked, gt, aliases, k)

            for name, val in [("Recall", r), ("Precision", p), ("F1", f),
                              ("HitRate", h), ("AP", ap), ("NDCG", nd)]:
                detail[f"{name}@{k}"] = f"{val:.4f}"
                accum.setdefault(f"{name}@{k}", []).append(val)

        detail_rows.append(detail)

    n_scored = len(detail_rows)
    summary = {"Scored_Questions": n_scored}
    for key, vals in accum.items():
        summary[f"Avg_{key}"] = sum(vals) / len(vals) if vals else 0.0

    return detail_rows, summary


# ── CSV/XLSX writers ─────────────────────────────────────────────────────────

def _detail_columns() -> list[str]:
    cols = ["No", "Pertanyaan", "GT_Total", "MRR"]
    for k in CUTOFFS:
        cols.append(f"Dok@{k}")
    for k in CUTOFFS:
        for m in ["Recall", "Precision", "F1", "HitRate", "AP", "NDCG"]:
            cols.append(f"{m}@{k}")
    return cols


def write_detail_csv(path, rows):
    cols = _detail_columns()
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_recap_csv(path, summary):
    rows_out = []
    rows_out.append({"Metric": "Scored_Questions", "Value": summary["Scored_Questions"]})

    mrr = summary.get("Avg_MRR", 0)
    rows_out.append({"Metric": "MRR", "Value": f"{mrr:.4f}"})

    for k in CUTOFFS:
        for m in ["Recall", "Precision", "F1", "HitRate", "AP", "NDCG"]:
            val = summary.get(f"Avg_{m}@{k}", 0)
            rows_out.append({"Metric": f"{m}@{k}", "Value": f"{val:.4f}"})

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["Metric", "Value"])
        writer.writeheader()
        writer.writerows(rows_out)


def write_xlsx(csv_path, xlsx_path):
    """Convert CSV to XLSX with comma decimals."""
    import openpyxl as xl
    def fmt(val):
        try:
            fv = float(val)
            s = f"{fv:.4f}".rstrip('0')
            if s.endswith('.'):
                s += '0'
            return s.replace('.', ',')
        except ValueError:
            return val

    wb = xl.Workbook()
    ws = wb.active
    with open(csv_path, newline='', encoding='utf-8') as f:
        reader = csv.reader(f)
        header = next(reader)
        ws.append(header)
        for row in reader:
            ws.append([fmt(cell) for cell in row])
    wb.save(xlsx_path)


def print_summary(label, summary):
    n = summary["Scored_Questions"]
    print(f"\n  {'='*60}")
    print(f"  {label}  ({n} scored)")
    print(f"  {'='*60}")
    print(f"  {'Metric':<16} {'Value':>10}")
    print(f"  {'─'*16} {'─'*10}")
    mrr = summary.get("Avg_MRR", 0)
    print(f"  {'MRR':<16} {mrr:>10.4f}")
    for k in CUTOFFS:
        print(f"  {'─'*16} {'─'*10}")
        for m in ["Recall", "Precision", "F1", "HitRate", "AP", "NDCG"]:
            v = summary.get(f"Avg_{m}@{k}", 0)
            print(f"  {m+'@'+str(k):<16} {v:>10.4f}")
    print(f"  {'='*60}")


# ── Main pipeline ────────────────────────────────────────────────────────────

def process_xlsx(xlsx_path: str):
    base = os.path.splitext(os.path.basename(xlsx_path))[0]
    qa_label = QA_LABELS.get(base, base)

    print(f"\n{'='*70}")
    print(f"  Benchmark v5: {qa_label}")
    print(f"  VDB_MAX_DOCS={VDB_MAX_DOCS}, RERANK_ALPHA={RERANK_ALPHA}")
    print(f"{'='*70}")

    # Connectivity checks
    if not pinecone_client.test_connection():
        print("ERROR: Pinecone not connected!")
        return
    print("  Pinecone ✓")

    print("  Testing HF embedding...")
    hf_ok = False
    for attempt in range(5):
        if llm_stance.test_hf_connection():
            hf_ok = True
            break
        print(f"    Attempt {attempt+1}/5 failed, retrying in 15s...")
        time.sleep(15)
    if not hf_ok:
        print("ERROR: HF embedding not connected!")
        return
    print("  HuggingFace ✓")

    try:
        neo4j_ok = neo4j_client.test_connection()
    except Exception:
        neo4j_ok = False
    print(f"  Neo4j {'✓' if neo4j_ok else '✗ (re-rank will use VDB order)'}")

    questions = _parse_questions(xlsx_path)
    print(f"  Questions: {len(questions)}\n")

    # ── Process each question ─────────────────────────────────────────────
    raw_results = []
    t0 = time.time()

    for idx, q in enumerate(questions):
        pct = (idx + 1) / len(questions) * 100
        short_q = q["question"][:65]
        print(f"  [{idx+1}/{len(questions)}] ({pct:.0f}%) {short_q}…", end="", flush=True)

        gt = q["gt_doc_ids"]
        q_injected = q["q_doc_ids"]

        if not gt:
            raw_results.append({
                "No": q["no"], "Pertanyaan": q["question"][:200],
                "GT_Total": 0, "GT_Doc_IDs": "", "GT_From_Question": "",
                "Dok_VDB": "", "Dok_GraphRAG": "",
            })
            print(" SKIP (no GT)")
            continue

        try:
            q_embedding = llm_stance.get_embedding(q["question"])

            # Step 1: VDB retrieval — get 20 unique docs WITH scores
            vdb_raw = pinecone_client.semantic_search(
                query_embedding=q_embedding, top_k=VDB_TOP_K
            )

            # Deduplicate, keep first occurrence (highest score per doc)
            seen = set()
            vdb_docs_with_scores = []
            for hit in vdb_raw:
                did = hit.get("doc_id", "")
                score = hit.get("score", 0.0)
                if did and did not in seen:
                    seen.add(did)
                    vdb_docs_with_scores.append((did, score))
                    if len(vdb_docs_with_scores) >= VDB_MAX_DOCS:
                        break

            vdb_doc_ids = [d[0] for d in vdb_docs_with_scores]

            # Step 2: Re-rank with Neo4j graph signal
            graphrag_doc_ids = rerank_with_graph(vdb_docs_with_scores, neo4j_ok)

            raw_results.append({
                "No": q["no"],
                "Pertanyaan": q["question"][:200],
                "GT_Total": len(gt),
                "GT_Doc_IDs": ", ".join(sorted(gt)),
                "GT_From_Question": ", ".join(sorted(q_injected)) if q_injected else "",
                "Dok_VDB": ", ".join(vdb_doc_ids),
                "Dok_GraphRAG": ", ".join(graphrag_doc_ids),
            })
            print(f" ✓ VDB={len(vdb_doc_ids)} docs, GraphRAG reranked")

        except Exception as e:
            raw_results.append({
                "No": q["no"], "Pertanyaan": q["question"][:200],
                "GT_Total": len(gt),
                "GT_Doc_IDs": ", ".join(sorted(gt)),
                "GT_From_Question": ", ".join(sorted(q_injected)) if q_injected else "",
                "Dok_VDB": f"Error: {e}", "Dok_GraphRAG": "",
            })
            print(f" ✗ {e}")

    elapsed = time.time() - t0
    print(f"\n  Retrieval done in {elapsed:.1f}s")

    # ── Write outputs with new naming convention ──────────────────────────
    os.makedirs(DETAIL_CSV_DIR, exist_ok=True)
    os.makedirs(DETAIL_XLSX_DIR, exist_ok=True)
    os.makedirs(RECAP_DIR, exist_ok=True)

    # Write raw detailed results (both methods in one file for reference)
    raw_cols = ["No", "Pertanyaan", "GT_Total", "GT_Doc_IDs", "GT_From_Question",
                "Dok_VDB", "Dok_GraphRAG"]
    raw_path = os.path.join(DETAIL_CSV_DIR, f"VDB+GraphRAG_ReRank-{qa_label}.csv")
    with open(raw_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=raw_cols, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(raw_results)
    print(f"  → {raw_path}")

    # ── Compute & write metrics per method ────────────────────────────────
    for method, doc_key in [("VDB", "Dok_VDB"), ("GraphRAG_ReRank", "Dok_GraphRAG")]:
        label = f"{method}-{qa_label}"

        detail_rows, summary = compute_single_method_metrics(raw_results, doc_key)

        # Detail CSV
        detail_csv = os.path.join(DETAIL_CSV_DIR, f"{label}-metrics.csv")
        write_detail_csv(detail_csv, detail_rows)
        print(f"  → {detail_csv}")

        # Detail XLSX
        detail_xlsx = os.path.join(DETAIL_XLSX_DIR, f"{label}-metrics.xlsx")
        write_xlsx(detail_csv, detail_xlsx)
        print(f"  → {detail_xlsx}")

        # Recap CSV
        recap_csv = os.path.join(RECAP_DIR, f"{label}.csv")
        write_recap_csv(recap_csv, summary)
        print(f"  → {recap_csv}")

        print_summary(label, summary)


# ── Entry point ──────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Benchmark v5 — VDB + Graph Re-Rank")
    parser.add_argument("xlsx", nargs="?", default=None)
    args = parser.parse_args()

    if args.xlsx:
        files = [args.xlsx]
    else:
        files = sorted(glob.glob(os.path.join(BENCHMARK_DIR, "*.xlsx")))
        if not files:
            print("No .xlsx files in benchmark/")
            sys.exit(1)

    print(f"Processing {len(files)} file(s)")
    for xlsx_path in files:
        process_xlsx(xlsx_path)

    print("\n✅ All done!")


if __name__ == "__main__":
    main()
