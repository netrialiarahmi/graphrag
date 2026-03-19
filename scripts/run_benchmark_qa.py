#!/usr/bin/env python3
"""
Benchmark QA — GraphRAG Pipeline Answer Evaluator

Reads all questions from an Excel benchmark file, runs each question through
the full GraphRAG search pipeline (same logic as app.py Tab 1), then uses
GPT to judge whether the generated answer is correct.

Output (Excel + CSV):
  output/qa_benchmark_<filename>.xlsx   — full table
  output/qa_benchmark_<filename>.csv    — same as CSV

Columns:
  No | Pertanyaan | Jawaban GraphRAG | Jawaban Benar | Verdict (BENAR/SALAH/PARSIAL)

Usage:
    python run_benchmark_qa.py
    python run_benchmark_qa.py path/to/file.xlsx
"""

import argparse
import csv
import os
import sys
import time

os.environ["GRAPHRAG_STANDALONE"] = "1"
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from dotenv import load_dotenv
load_dotenv()

from utils import llm_stance, neo4j_client, pinecone_client
from utils.benchmark_helpers import (
    get_unique_doc_ids,
    extract_doc_ids_from_question,
)

# ── Constants ─────────────────────────────────────────────────────────────────
ROOT_DIR      = os.path.dirname(os.path.abspath(__file__))
BENCHMARK_DIR = os.path.join(ROOT_DIR, "benchmark")
OUTPUT_DIR    = os.path.join(ROOT_DIR, "output")
DEFAULT_FILE  = os.path.join(BENCHMARK_DIR, "govnetic_qa_complete_50 (business).xlsx")

DELAY_BETWEEN_Q = 2  # seconds between questions (avoid rate-limit)


# ── Pipeline helpers (mirrors app.py) ─────────────────────────────────────────

def _build_interleaved_context(
    primary_doc_ids: list,
    related_doc_ids: list,
    context_docs: dict,
    max_chunks: int = 30,
    max_chars: int = 12000,
) -> list:
    result = []
    total_chars = 0
    doc_queues = {}
    for did in primary_doc_ids + related_doc_ids:
        info = context_docs.get(did)
        if not info:
            continue
        chunks = list(info["chunks"])
        scored = sorted(
            [c for c in chunks if c.get("score") is not None],
            key=lambda c: c.get("score", 0), reverse=True,
        )
        unscored = [c for c in chunks if c.get("score") is None]
        doc_queues[did] = scored + unscored

    if not doc_queues:
        return []

    seen_ids = set()
    doc_keys = list(doc_queues.keys())
    idx_map = {did: 0 for did in doc_keys}
    exhausted = set()

    while (
        len(result) < max_chunks
        and total_chars < max_chars
        and len(exhausted) < len(doc_keys)
    ):
        for did in doc_keys:
            if did in exhausted:
                continue
            queue = doc_queues[did]
            idx = idx_map[did]
            if idx >= len(queue):
                exhausted.add(did)
                continue
            chunk = queue[idx]
            idx_map[did] = idx + 1
            cid = chunk.get("id", "")
            if cid and cid in seen_ids:
                continue
            if cid:
                seen_ids.add(cid)
            content = chunk.get("content", "")
            total_chars += len(content)
            result.append(chunk)
            if len(result) >= max_chunks or total_chars >= max_chars:
                break

    return result


def run_pipeline(query: str, neo4j_ok: bool) -> tuple[str, list[str]]:
    """Run the full GraphRAG search pipeline for a single query.

    Returns:
        (answer_text, primary_doc_ids)
    """
    # ── Phase A: Query Analysis ───────────────────────────────────────────────
    regex_doc_ids = extract_doc_ids_from_question(query)
    expanded_terms = llm_stance.expand_query(query)

    # ── Phase B: GraphRAG Document Discovery ─────────────────────────────────
    graph_doc_ids = []
    if neo4j_ok:
        all_neo4j_docs = neo4j_client.get_all_documents()
        graph_doc_ids = llm_stance.smart_doc_lookup(query, all_neo4j_docs)

    # ── Gate 1: Sufficiency check (regex + GraphRAG) ──────────────────────────
    gate1_doc_ids = list(dict.fromkeys(list(graph_doc_ids) + list(regex_doc_ids)))

    early_exit = False
    raw_results = []
    semantic_chunks_by_doc = {}
    graph_expanded_ids = []
    primary_doc_ids = []
    merged_doc_ids = []
    added = set()

    if gate1_doc_ids and neo4j_ok:
        gate1_summaries = {}
        for did in gate1_doc_ids[:10]:
            try:
                detail = neo4j_client.get_document_detail(did)
                pasals = detail.get("pasals", [])
                if pasals:
                    texts = [p.get("content", "") or p.get("name", "") for p in pasals[:3]]
                    gate1_summaries[did] = " ".join(t for t in texts if t)[:500]
                elif detail.get("document", {}).get("judul"):
                    gate1_summaries[did] = detail["document"]["judul"]
                else:
                    gate1_summaries[did] = ""
            except Exception:
                gate1_summaries[did] = ""

        if llm_stance.judge_sufficiency(query, gate1_doc_ids, gate1_summaries):
            early_exit = True
            primary_doc_ids = gate1_doc_ids[:7]
            merged_doc_ids = gate1_doc_ids
            added = set(gate1_doc_ids)

    if not early_exit:
        # ── Phase C (lite): VDB Semantic Search ──────────────────────────────
        query_embedding = llm_stance.get_embedding(query)
        raw_results = pinecone_client.semantic_search(query_embedding=query_embedding, top_k=30)

        for hit in raw_results:
            did = hit.get("doc_id", "")
            if did:
                semantic_chunks_by_doc.setdefault(did, []).append(hit)

        vdb_doc_ids = get_unique_doc_ids(raw_results, 10)

        # ── Phase D: Merge ────────────────────────────────────────────────────
        for did in graph_doc_ids:
            if did not in added:
                merged_doc_ids.append(did)
                added.add(did)
        for did in regex_doc_ids:
            if did not in added:
                merged_doc_ids.append(did)
                added.add(did)
        for did in vdb_doc_ids:
            if did not in added:
                merged_doc_ids.append(did)
                added.add(did)

        if not merged_doc_ids:
            return "(Tidak ditemukan dokumen relevan.)", []

        # ── Gate 2: Sufficiency check (merged candidates) ─────────────────────
        gate2_summaries = {}
        for did in merged_doc_ids[:10]:
            if did in semantic_chunks_by_doc:
                gate2_summaries[did] = semantic_chunks_by_doc[did][0].get("content", "")[:500]
            else:
                gate2_summaries[did] = ""

        gate2_passed = llm_stance.judge_sufficiency(query, merged_doc_ids, gate2_summaries)

        if gate2_passed:
            primary_doc_ids = merged_doc_ids[:7]
            for did in list(graph_doc_ids[:5]) + list(regex_doc_ids):
                if did not in primary_doc_ids:
                    primary_doc_ids.append(did)
        else:
            # ── Phase C (full): Expand VDB search ────────────────────────────
            if len(raw_results) < 60:
                more_results = pinecone_client.semantic_search(query_embedding=query_embedding, top_k=100)
                seen_ids = {h.get("id", "") for h in raw_results}
                for hit in more_results:
                    hid = hit.get("id", "")
                    if hid and hid not in seen_ids:
                        raw_results.append(hit)
                        seen_ids.add(hid)
                        did = hit.get("doc_id", "")
                        if did:
                            semantic_chunks_by_doc.setdefault(did, []).append(hit)

            seen_ids_exp = {h.get("id", "") for h in raw_results}
            for term in expanded_terms[:3]:
                try:
                    term_emb = llm_stance.get_embedding(term)
                    term_results = pinecone_client.semantic_search(query_embedding=term_emb, top_k=30)
                    for hit in term_results:
                        hid = hit.get("id", "")
                        if hid and hid not in seen_ids_exp:
                            raw_results.append(hit)
                            seen_ids_exp.add(hid)
                            did = hit.get("doc_id", "")
                            if did:
                                semantic_chunks_by_doc.setdefault(did, []).append(hit)
                except Exception:
                    pass

            vdb_doc_ids_full = get_unique_doc_ids(raw_results, 20)
            for did in vdb_doc_ids_full:
                if did not in added:
                    merged_doc_ids.append(did)
                    added.add(did)

            # ── Phase E: Deep Graph Traversal ─────────────────────────────────
            if neo4j_ok:
                for did in merged_doc_ids[:5]:
                    try:
                        subgraph = neo4j_client.get_citing_documents(did, hops=2)
                        for node in subgraph.get("nodes", []):
                            ndid = node.get("doc_id", "")
                            if ndid and ndid not in added:
                                graph_expanded_ids.append(ndid)
                                added.add(ndid)
                    except Exception:
                        pass

            # ── Phase F: Re-rank ──────────────────────────────────────────────
            all_candidate_ids = merged_doc_ids + graph_expanded_ids
            doc_summaries = {}
            for did in all_candidate_ids:
                summary = ""
                if did in semantic_chunks_by_doc:
                    summary = semantic_chunks_by_doc[did][0].get("content", "")
                if not summary:
                    vdb_chunks = pinecone_client.fetch_by_doc_id(did, top_k=3)
                    if vdb_chunks:
                        summary = vdb_chunks[0].get("content", "")
                if not summary and neo4j_ok:
                    try:
                        detail = neo4j_client.get_document_detail(did)
                        pasals = detail.get("pasals", [])
                        if pasals:
                            texts = [p.get("content", "") or p.get("name", "") for p in pasals[:3]]
                            summary = " ".join(t for t in texts if t)
                    except Exception:
                        pass
                doc_summaries[did] = summary[:500] if summary else ""

            ranked = llm_stance.rerank_documents(query, doc_summaries)
            primary_doc_ids = [did for did, score in ranked if score >= 3.0][:7]
            for did in list(graph_doc_ids[:5]) + list(regex_doc_ids):
                if did not in primary_doc_ids:
                    primary_doc_ids.append(did)
            if not primary_doc_ids:
                primary_doc_ids = merged_doc_ids[:5]

    # ── Phase G: Content Assembly ─────────────────────────────────────────────
    context_docs = {}
    seen_chunk_ids = set()

    for did in primary_doc_ids:
        doc_chunks = []
        sem_chunks = semantic_chunks_by_doc.get(did, [])
        for ch in sem_chunks:
            doc_chunks.append(ch)
            seen_chunk_ids.add(ch.get("id", ""))

        extra = pinecone_client.fetch_by_doc_id(did, top_k=80)
        for ch in extra:
            cid = ch.get("id", "")
            if cid not in seen_chunk_ids:
                doc_chunks.append(ch)
                seen_chunk_ids.add(cid)

        # G3. Neo4j Pasal/Ayat content (always for primary docs)
        if neo4j_ok:
            try:
                detail = neo4j_client.get_document_detail(did)
                for pasal in detail.get("pasals", []):
                    p_content = pasal.get("content", "")
                    p_name = pasal.get("name", "")
                    neo4j_id = f"neo4j-{did}-{p_name}"
                    if p_content and len(p_content) > 20 and neo4j_id not in seen_chunk_ids:
                        doc_chunks.append({
                            "id": neo4j_id,
                            "doc_id": did, "article_id": p_name,
                            "content": p_content, "scope": "neo4j-pasal",
                        })
                        seen_chunk_ids.add(neo4j_id)
                for ayat in detail.get("ayats", []):
                    a_content = ayat.get("content", "")
                    a_name = ayat.get("name", "")
                    p_name = ayat.get("pasal_name", "")
                    neo4j_id = f"neo4j-{did}-{p_name}-{a_name}"
                    if a_content and len(a_content) > 20 and neo4j_id not in seen_chunk_ids:
                        doc_chunks.append({
                            "id": neo4j_id,
                            "doc_id": did, "article_id": f"{p_name} {a_name}",
                            "content": a_content, "scope": "neo4j-ayat",
                        })
                        seen_chunk_ids.add(neo4j_id)
            except Exception:
                pass

        context_docs[did] = {"source": "pipeline", "chunks": doc_chunks}

    # G4. Neo4j direct neighbors as supplementary
    related_doc_ids = []
    if neo4j_ok:
        for did in primary_doc_ids[:3]:
            related = neo4j_client.get_related_documents(did, limit=2)
            for rdoc in related:
                rdid = rdoc.get("doc_id", "")
                if rdid and rdid not in context_docs and rdid not in related_doc_ids:
                    related_doc_ids.append(rdid)

    for rdid in related_doc_ids:
        chunks = pinecone_client.fetch_by_doc_id(rdid, top_k=30)
        if chunks:
            context_docs[rdid] = {"source": "Neo4j (Related)", "chunks": chunks}

    # ── Phase H: Answer Generation ────────────────────────────────────────────
    # Build relationship context from Neo4j graph edges
    relationship_context = ""
    if neo4j_ok:
        try:
            all_doc_ids = list(context_docs.keys())
            edges_data = neo4j_client.get_edges_between(all_doc_ids)
            rel_lines = []
            for edge in edges_data.get("edges", []):
                src = edge.get("source_id", "")
                tgt = edge.get("target_id", "")
                rel_type = edge.get("type", "")
                if src and tgt and rel_type:
                    rel_lines.append(f"- {src} --[{rel_type}]--> {tgt}")
            if rel_lines:
                relationship_context = "\n".join(rel_lines)
        except Exception:
            pass

    llm_chunks = _build_interleaved_context(
        primary_doc_ids=primary_doc_ids,
        related_doc_ids=related_doc_ids,
        context_docs=context_docs,
        max_chunks=40,
        max_chars=16000,
    )

    answer = llm_stance.ask_about_documents(
        query, llm_chunks,
        relationship_context=relationship_context,
    )
    return answer, primary_doc_ids


def judge_answer(query: str, generated_answer: str, expected_answer: str) -> str:
    """Use LLM to judge if the generated answer is correct vs expected.

    Returns: 'BENAR', 'PARSIAL', or 'SALAH'
    """
    client = llm_stance.get_llm_client()
    prompt = f"""Kamu adalah evaluator jawaban hukum Indonesia. Bandingkan JAWABAN SISTEM dengan JAWABAN BENAR.

PERTANYAAN:
{query}

JAWABAN SISTEM:
{generated_answer}

JAWABAN BENAR (referensi):
{expected_answer}

Evaluasi apakah JAWABAN SISTEM sudah benar, parsial, atau salah dibandingkan JAWABAN BENAR.
- BENAR: jawaban sistem akurat dan mencakup poin utama dari jawaban benar
- PARSIAL: jawaban sistem benar sebagian namun ada informasi penting yang kurang atau sedikit keliru
- SALAH: jawaban sistem salah atau bertolak belakang dengan jawaban benar

Balas HANYA dengan satu kata: BENAR, PARSIAL, atau SALAH."""

    try:
        resp = client.chat.completions.create(
            model=llm_stance.LLM_MODEL,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=16,
            temperature=0.0,
        )
        verdict = resp.choices[0].message.content.strip().upper()
        if "BENAR" in verdict and "SALAH" not in verdict:
            return "BENAR"
        elif "PARSIAL" in verdict:
            return "PARSIAL"
        elif "SALAH" in verdict:
            return "SALAH"
        else:
            return verdict[:20]
    except Exception as e:
        return f"ERROR: {e}"


def _parse_questions(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    questions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        q_no    = str(row[0])
        q_text  = str(row[1]) if row[1] else ""
        q_answer = str(row[2]) if len(row) > 2 and row[2] else ""
        if q_text:
            questions.append({"no": q_no, "question": q_text, "expected": q_answer})
    wb.close()
    return questions


def _save_results(results: list[dict], xlsx_path: str, base_name: str):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_csv  = os.path.join(OUTPUT_DIR, f"qa_benchmark_{base_name}.csv")
    out_xlsx = os.path.join(OUTPUT_DIR, f"qa_benchmark_{base_name}.xlsx")

    # CSV
    with open(out_csv, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=["No", "Pertanyaan", "Jawaban GraphRAG", "Jawaban Benar", "Verdict"])
        writer.writeheader()
        writer.writerows(results)

    # Excel with auto-width columns
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Benchmark QA"

    headers = ["No", "Pertanyaan", "Jawaban GraphRAG", "Jawaban Benar", "Verdict"]
    # Header style
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Verdict colors
    verdict_colors = {"BENAR": "D1FAE5", "PARSIAL": "FEF9C3", "SALAH": "FEE2E2"}

    for row_data in results:
        ws.append([
            row_data["No"],
            row_data["Pertanyaan"],
            row_data["Jawaban GraphRAG"],
            row_data["Jawaban Benar"],
            row_data["Verdict"],
        ])
        # Color verdict cell
        verdict_val = row_data["Verdict"]
        color = verdict_colors.get(verdict_val, "FFFFFF")
        verdict_cell = ws.cell(row=ws.max_row, column=5)
        verdict_cell.fill = PatternFill("solid", fgColor=color)
        verdict_cell.font = Font(bold=True)
        verdict_cell.alignment = Alignment(horizontal="center")

    # Column widths
    col_widths = [10, 50, 70, 70, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Wrap text for all data cells
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(out_xlsx)
    return out_csv, out_xlsx


def main():
    parser = argparse.ArgumentParser(description="GraphRAG QA Benchmark Runner")
    parser.add_argument("xlsx", nargs="?", default=DEFAULT_FILE, help="Path to benchmark Excel file")
    args = parser.parse_args()

    xlsx_path = args.xlsx
    if not os.path.exists(xlsx_path):
        print(f"[ERROR] File not found: {xlsx_path}")
        sys.exit(1)

    base_name = os.path.splitext(os.path.basename(xlsx_path))[0].replace(" ", "_")
    print(f"\n{'='*60}")
    print(f"  GraphRAG QA Benchmark")
    print(f"  File : {os.path.basename(xlsx_path)}")
    print(f"{'='*60}\n")

    # Check Neo4j connectivity
    neo4j_ok = False
    try:
        neo4j_client.get_all_documents()
        neo4j_ok = True
        print("[OK] Neo4j connected")
    except Exception as e:
        print(f"[WARN] Neo4j not available: {e}")

    questions = _parse_questions(xlsx_path)
    print(f"[INFO] Loaded {len(questions)} questions\n")

    results = []
    # Checkpoint: load partial results if run was interrupted
    out_csv = os.path.join(OUTPUT_DIR, f"qa_benchmark_{base_name}.csv")
    done_nos = set()
    if os.path.exists(out_csv):
        with open(out_csv, newline="", encoding="utf-8-sig") as f:
            for r in csv.DictReader(f):
                # Only skip if verdict is a valid result (not an error)
                if r["Verdict"] in ("BENAR", "PARSIAL", "SALAH"):
                    done_nos.add(r["No"])
                results.append(r)
        print(f"[INFO] Resuming — {len(done_nos)} questions with valid verdict, {len(results)} total loaded\n")

    for i, q in enumerate(questions):
        if q["no"] in done_nos:
            print(f"[{i+1:02d}/{len(questions)}] {q['no']} — skipped (already done)")
            continue

        print(f"[{i+1:02d}/{len(questions)}] {q['no']} — {q['question'][:80]}…")

        # Check if we already have an answer (partial result with error verdict)
        existing = next((r for r in results if r["No"] == q["no"]), None)
        if existing and existing.get("Jawaban GraphRAG") and not existing["Jawaban GraphRAG"].startswith("(Pipeline error"):
            # Re-use existing answer, only re-judge
            answer = existing["Jawaban GraphRAG"]
            print(f"          ↳ re-using cached answer, re-judging verdict…")
            results = [r for r in results if r["No"] != q["no"]]  # remove old row
        else:
            try:
                answer, doc_ids = run_pipeline(q["question"], neo4j_ok)
                print(f"          ↳ docs: {doc_ids[:5]}")
            except Exception as e:
                answer = f"(Pipeline error: {e})"
                print(f"          ↳ ERROR: {e}")

        try:
            verdict = judge_answer(q["question"], answer, q["expected"])
            print(f"          ↳ verdict: {verdict}")
        except Exception as e:
            verdict = f"ERROR: {e}"

        results.append({
            "No": q["no"],
            "Pertanyaan": q["question"],
            "Jawaban GraphRAG": answer,
            "Jawaban Benar": q["expected"],
            "Verdict": verdict,
        })

        # Save checkpoint after each question
        _save_results(results, xlsx_path, base_name)

        if i < len(questions) - 1:
            time.sleep(DELAY_BETWEEN_Q)

    out_csv, out_xlsx = _save_results(results, xlsx_path, base_name)

    # Summary
    verdicts = [r["Verdict"] for r in results]
    benar   = verdicts.count("BENAR")
    parsial = verdicts.count("PARSIAL")
    salah   = verdicts.count("SALAH")

    print(f"\n{'='*60}")
    print(f"  HASIL BENCHMARK")
    print(f"{'='*60}")
    print(f"  Total  : {len(results)}")
    print(f"  BENAR  : {benar}  ({benar/len(results)*100:.1f}%)")
    print(f"  PARSIAL: {parsial}  ({parsial/len(results)*100:.1f}%)")
    print(f"  SALAH  : {salah}  ({salah/len(results)*100:.1f}%)")
    print(f"\n  Output CSV  : {out_csv}")
    print(f"  Output XLSX : {out_xlsx}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
