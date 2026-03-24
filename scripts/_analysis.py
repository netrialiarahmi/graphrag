"""Comprehensive analysis of semantic node retrieval capability."""
import csv, re, sys, os
from collections import Counter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

DOC_ALIASES = {
    'UU 40/2007': 'UU-NASIONAL-40-2007',
    'UU No. 40 Tahun 2007': 'UU-NASIONAL-40-2007',
    'UU 11/2020': 'UU-NASIONAL-11-2020',
    'UU No. 11 Tahun 2020': 'UU-NASIONAL-11-2020',
    'Perppu 2/2022': 'PERPPU-NASIONAL-2-2022',
    'Perppu No. 2 Tahun 2022': 'PERPPU-NASIONAL-2-2022',
    'PP 16/2021': 'PP-NASIONAL-16-2021',
    'PP No. 16 Tahun 2021': 'PP-NASIONAL-16-2021',
    'UU 23/2014': 'UU-NASIONAL-23-2014',
    'UU No. 23 Tahun 2014': 'UU-NASIONAL-23-2014',
    'UU 5/1999': 'UU-NASIONAL-5-1999',
    'UU No. 5 Tahun 1999': 'UU-NASIONAL-5-1999',
    'UU 12/2011': 'UU-NASIONAL-12-2011',
    'UU No. 12 Tahun 2011': 'UU-NASIONAL-12-2011',
    'UU 28/2002': 'UU-NASIONAL-28-2002',
    'UU No. 28 Tahun 2002': 'UU-NASIONAL-28-2002',
    'UU 20/2008': 'UU-NASIONAL-20-2008',
    'UU No. 20 Tahun 2008': 'UU-NASIONAL-20-2008',
    'UU 11/2014': 'UU-NASIONAL-11-2014',
    'UU 31/2002': 'UU-NASIONAL-31-2002',
    'PP 29/2021': 'PP-NASIONAL-29-2021',
    'UU 13/2022': 'UU-NASIONAL-13-2022',
}

import openpyxl
wb = openpyxl.load_workbook('benchmark/govnetic_qa_complete_50 (business).xlsx')
ws = wb.active
headers = [c.value for c in ws[1]]
eidx = headers.index('Evidence (Pasal/Ayat)')
qidx = headers.index('Pertanyaan')
catidx = headers.index('Kategori Pertanyaan')
compidx = headers.index('Complexity')

def parse_expected(ev):
    ids = set()
    if not ev:
        return ids
    pats = re.findall(
        r'(?:UU|PP|Perppu|Permen\w*|Kepmen\w*|Pergub|Perda)\s*(?:No\.\s*)?\d+(?:/\d+)?(?:\s+Tahun\s+\d+)?',
        str(ev), re.IGNORECASE
    )
    for p in pats:
        p = p.strip()
        if p in DOC_ALIASES:
            ids.add(DOC_ALIASES[p])
        else:
            m = re.search(r'(\d+)[/\s]+(?:Tahun\s+)?(\d{4})', p)
            if m:
                num, yr = m.group(1), m.group(2)
                prefix = p.split()[0].upper()
                if prefix == 'PERPPU':
                    did = f'PERPPU-NASIONAL-{num}-{yr}'
                elif prefix == 'PP':
                    did = f'PP-NASIONAL-{num}-{yr}'
                elif prefix == 'UU':
                    did = f'UU-NASIONAL-{num}-{yr}'
                else:
                    did = f'{prefix}-NASIONAL-{num}-{yr}'
                ids.add(did)
    return ids


bench = 'output/qa_benchmark_semantic_govnetic_qa_complete_50_(business).csv'
with open(bench, newline='', encoding='utf-8-sig') as f:
    results = list(csv.DictReader(f))

# ── 1. Verdict Summary ──
verdicts = Counter(r['Verdict'] for r in results)
print("=" * 70)
print("1. LLM VERDICT SUMMARY (potentially biased)")
print("=" * 70)
for v in ['BENAR', 'PARSIAL', 'SALAH']:
    c = verdicts.get(v, 0)
    print(f"   {v:8s}: {c:3d} / {len(results)}  ({c/len(results)*100:.1f}%)")

# ── 2. Per-question retrieval hit/miss ──
expected_counts = Counter()
retrieved_hits = Counter()
question_status = []  # (tc_id, question, verdict, retrieval_status, expected, retrieved, missing)

for i, row in enumerate(results):
    ws_row = ws[i + 2]
    ev = ws_row[eidx].value
    cat = ws_row[catidx].value or ''
    comp = ws_row[compidx].value or ''
    expected = parse_expected(ev)
    docs_retrieved = set(d.strip() for d in row.get('Docs', '').split(',') if d.strip())
    
    found = expected & docs_retrieved
    missing = expected - docs_retrieved
    
    if not expected:
        status = 'NO_EXPECTED'
    elif missing == set():
        status = 'FULL_HIT'
    elif found:
        status = 'PARTIAL'
    else:
        status = 'MISS'
    
    for e in expected:
        expected_counts[e] += 1
        if e in docs_retrieved:
            retrieved_hits[e] += 1
    
    question_status.append({
        'no': row.get('No', i+1),
        'question': row.get('Pertanyaan', ''),
        'verdict': row.get('Verdict', ''),
        'status': status,
        'expected': expected,
        'retrieved': docs_retrieved,
        'missing': missing,
        'category': cat,
        'complexity': comp,
    })

print()
print("=" * 70)
print("2. RETRIEVAL ACCURACY (document-level)")
print("=" * 70)
status_counts = Counter(q['status'] for q in question_status)
for s in ['FULL_HIT', 'PARTIAL', 'MISS', 'NO_EXPECTED']:
    c = status_counts.get(s, 0)
    print(f"   {s:12s}: {c:3d} / {len(results)}  ({c/len(results)*100:.1f}%)")

total_exp = sum(expected_counts.values())
total_found = sum(retrieved_hits.values())
print(f"\n   Document Recall: {total_found}/{total_exp} = {total_found/total_exp*100:.1f}%")

# ── 3. Per-document retrieval success rate ──
print()
print("=" * 70)
print("3. PER-DOCUMENT RETRIEVAL SUCCESS RATE")
print("=" * 70)
print(f"   {'Document':<42s} Expected  Found   Rate")
print("   " + "-" * 65)
for doc, cnt in expected_counts.most_common():
    found = retrieved_hits.get(doc, 0)
    rate = found / cnt * 100
    marker = " ⚠️" if rate == 0 else ""
    print(f"   {doc:<42s}  {cnt:5d}  {found:5d}  {rate:5.1f}%{marker}")

# ── 4. Most retrieved documents (what search actually returns) ──
all_retrieved = Counter()
for r in results:
    for d in (d.strip() for d in r.get('Docs', '').split(',') if d.strip()):
        all_retrieved[d] += 1

print()
print("=" * 70)
print("4. MOST RETRIEVED DOCUMENTS (what search actually returns)")
print("=" * 70)
for d, c in all_retrieved.most_common(10):
    expected_flag = " (expected)" if d in expected_counts else " (noise)"
    print(f"   {c:3d}x  {d}{expected_flag}")

# ── 5. Cross-tab: verdict × retrieval  ──
print()
print("=" * 70)
print("5. CROSS-TAB: LLM Verdict × Retrieval Status")
print("=" * 70)
cross = Counter()
for q in question_status:
    cross[(q['verdict'], q['status'])] += 1

print(f"   {'':12s} FULL_HIT  PARTIAL  MISS  NO_EXP")
for v in ['BENAR', 'PARSIAL', 'SALAH']:
    row = f"   {v:12s}"
    for s in ['FULL_HIT', 'PARTIAL', 'MISS', 'NO_EXPECTED']:
        row += f"  {cross.get((v,s),0):7d}"
    print(row)

# ── 6. SUSPICIOUS: BENAR but bad retrieval ──
suspicious = [q for q in question_status if q['verdict'] == 'BENAR' and q['status'] in ('PARTIAL', 'MISS')]
print()
print("=" * 70)
print(f"6. SUSPICIOUS CASES: BENAR verdict but PARTIAL/MISS retrieval ({len(suspicious)} cases)")
print("=" * 70)
for q in suspicious:
    print(f"   [{q['no']}] {q['question'][:70]}")
    print(f"       Status: {q['status']}  |  Missing: {q['missing']}")

# ── 7. Category breakdown ──
print()
print("=" * 70)
print("7. RETRIEVAL BY QUESTION CATEGORY")
print("=" * 70)
cat_stats = {}
for q in question_status:
    cat = q['category'] or 'Unknown'
    if cat not in cat_stats:
        cat_stats[cat] = {'total': 0, 'full': 0, 'partial': 0, 'miss': 0, 'benar': 0}
    cat_stats[cat]['total'] += 1
    if q['status'] == 'FULL_HIT':
        cat_stats[cat]['full'] += 1
    elif q['status'] == 'PARTIAL':
        cat_stats[cat]['partial'] += 1
    elif q['status'] == 'MISS':
        cat_stats[cat]['miss'] += 1
    if q['verdict'] == 'BENAR':
        cat_stats[cat]['benar'] += 1

for cat, s in sorted(cat_stats.items(), key=lambda x: -x[1]['total']):
    print(f"   {cat} (n={s['total']})")
    print(f"     Retrieval: FULL {s['full']} | PARTIAL {s['partial']} | MISS {s['miss']}  =>  Full Hit Rate: {s['full']/s['total']*100:.0f}%")
    print(f"     LLM BENAR: {s['benar']}/{s['total']} ({s['benar']/s['total']*100:.0f}%)")

# ── 8. Complexity breakdown ──
print()
print("=" * 70)
print("8. RETRIEVAL BY COMPLEXITY")
print("=" * 70)
comp_stats = {}
for q in question_status:
    comp = q['complexity'] or 'Unknown'
    if comp not in comp_stats:
        comp_stats[comp] = {'total': 0, 'full': 0, 'partial': 0, 'miss': 0, 'benar': 0}
    comp_stats[comp]['total'] += 1
    if q['status'] == 'FULL_HIT':
        comp_stats[comp]['full'] += 1
    elif q['status'] == 'PARTIAL':
        comp_stats[comp]['partial'] += 1
    elif q['status'] == 'MISS':
        comp_stats[comp]['miss'] += 1
    if q['verdict'] == 'BENAR':
        comp_stats[comp]['benar'] += 1

for comp, s in sorted(comp_stats.items()):
    print(f"   {comp} (n={s['total']})")
    print(f"     Retrieval: FULL {s['full']} | PARTIAL {s['partial']} | MISS {s['miss']}  =>  Full Hit Rate: {s['full']/s['total']*100:.0f}%")
    print(f"     LLM BENAR: {s['benar']}/{s['total']} ({s['benar']/s['total']*100:.0f}%)")

# ── 9. Final verdict: how much can semantic carry alone? ──
genuine_benar = len([q for q in question_status if q['verdict'] == 'BENAR' and q['status'] == 'FULL_HIT'])
partial_benar = len([q for q in question_status if q['verdict'] == 'BENAR' and q['status'] == 'PARTIAL'])
usable = len([q for q in question_status if q['status'] in ('FULL_HIT', 'PARTIAL')])

print()
print("=" * 70)
print("9. FINAL VERDICT: SEMANTIC NODE SOLO CAPABILITY")
print("=" * 70)
print(f"   Genuinely correct (BENAR + FULL_HIT):  {genuine_benar}/{len(results)} = {genuine_benar/len(results)*100:.1f}%")
print(f"   Retrieval usable (FULL + PARTIAL):      {usable}/{len(results)} = {usable/len(results)*100:.1f}%")
print(f"   Total retrieval MISS:                   {status_counts.get('MISS',0)}/{len(results)} = {status_counts.get('MISS',0)/len(results)*100:.1f}%")
print(f"   Document Recall:                        {total_found}/{total_exp} = {total_found/total_exp*100:.1f}%")
print(f"   LLM Knowledge Contamination:            {len(suspicious)}/{verdicts.get('BENAR',0)} BENAR are suspicious")
print()
print("   CONCLUSION:")
print(f"   - Semantic node can RELIABLY handle ~{genuine_benar/len(results)*100:.0f}% of questions alone (FULL_HIT + BENAR)")
print(f"   - With partial evidence, it can ATTEMPT ~{usable/len(results)*100:.0f}% (but needs graph/deep for completeness)")  
print(f"   - {status_counts.get('MISS',0)/len(results)*100:.0f}% of questions get ZERO useful documents — graph/deep is essential")
print(f"   - {len(suspicious)}/{verdicts.get('BENAR',0)} 'correct' answers are from LLM knowledge, NOT from retrieved chunks")
print(f"   - Biggest gap: UU-NASIONAL-40-2007 (expected {expected_counts.get('UU-NASIONAL-40-2007',0)}x, found {retrieved_hits.get('UU-NASIONAL-40-2007',0)}x)")
print(f"   - 'Attractor' docs: UU-11-2020 appears {all_retrieved.get('UU-NASIONAL-11-2020',0)}x, PERPPU-2-2022 appears {all_retrieved.get('PERPPU-NASIONAL-2-2022',0)}x")
